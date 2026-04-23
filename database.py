"""
database.py - Backend SQLite persistente do sistema.
Arquivo salvo em: C:/Users/Admin/Documents/Robson/sp_indicadores.db
"""

import sqlite3
import os
from datetime import datetime

# Banco ao lado da pasta sp_dashboard (na raiz do projeto)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.normpath(os.path.join(BASE_DIR, "..", "sp_indicadores.db"))

SCHEMA = """
CREATE TABLE IF NOT EXISTS indicadores_mapeamento (
    id                       INTEGER PRIMARY KEY AUTOINCREMENT,
    codigo_indicador         TEXT    UNIQUE NOT NULL,
    nome_indicador           TEXT    NOT NULL,
    usa_dados_operacionais   INTEGER DEFAULT 1,
    aba_origem_excel         TEXT,
    campo_origem             TEXT,
    resultado_representa     TEXT,
    subindicadores_existem   INTEGER DEFAULT 0,
    subindicadores_status    TEXT    DEFAULT 'A definir',
    status_mapeamento        TEXT    DEFAULT 'Sem vínculo',
    observacoes              TEXT,
    indicador_ativo          INTEGER DEFAULT 1,
    modo_comparacao          TEXT    DEFAULT '2025 x 2026',
    atualizado_em            TEXT
);

CREATE TABLE IF NOT EXISTS importacoes_log (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    data_hora  TEXT,
    arquivo    TEXT,
    registros  INTEGER,
    status     TEXT,
    mensagem   TEXT
);
"""


def _connect():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Cria o banco e as tabelas se não existirem."""
    conn = _connect()
    try:
        for stmt in SCHEMA.strip().split(";"):
            s = stmt.strip()
            if s:
                conn.execute(s)
        conn.commit()
    finally:
        conn.close()


def is_empty() -> bool:
    conn = _connect()
    try:
        row = conn.execute("SELECT COUNT(*) FROM indicadores_mapeamento").fetchone()
        return row[0] == 0
    finally:
        conn.close()


def seed_from_mapping():
    """Popula o banco a partir dos dados do mapping_db.py (só se estiver vazio)."""
    from mapping_db import MAPEAMENTO_INDICADORES
    conn = _connect()
    try:
        now = datetime.now().isoformat(sep=" ", timespec="seconds")
        for m in MAPEAMENTO_INDICADORES:
            conn.execute("""
                INSERT OR IGNORE INTO indicadores_mapeamento
                    (codigo_indicador, nome_indicador, usa_dados_operacionais,
                     aba_origem_excel, campo_origem, resultado_representa,
                     subindicadores_existem, subindicadores_status,
                     status_mapeamento, observacoes, atualizado_em)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, (
                m["codigo_indicador"],
                m["nome_indicador"],
                1 if m["usa_dados_operacionais"] else 0,
                m.get("aba_origem_excel"),
                m.get("campo_origem"),
                m.get("resultado_representa"),
                1 if m.get("subindicadores_existem") else 0,
                m.get("subindicadores_status", "A definir"),
                m.get("status_mapeamento", "Sem vínculo"),
                m.get("observacoes"),
                now,
            ))
        conn.commit()
    finally:
        conn.close()


def get_all() -> list[dict]:
    """Retorna todos os registros como lista de dicts."""
    conn = _connect()
    try:
        rows = conn.execute(
            "SELECT * FROM indicadores_mapeamento ORDER BY codigo_indicador"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def get_by_codigo(codigo: str) -> dict | None:
    conn = _connect()
    try:
        row = conn.execute(
            "SELECT * FROM indicadores_mapeamento WHERE codigo_indicador = ?", (codigo,)
        ).fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


def upsert(record: dict) -> bool:
    """Insere ou atualiza um registro. Retorna True se bem-sucedido."""
    conn = _connect()
    try:
        record["atualizado_em"] = datetime.now().isoformat(sep=" ", timespec="seconds")
        conn.execute("""
            INSERT INTO indicadores_mapeamento
                (codigo_indicador, nome_indicador, usa_dados_operacionais,
                 aba_origem_excel, campo_origem, resultado_representa,
                 subindicadores_existem, subindicadores_status,
                 status_mapeamento, observacoes, indicador_ativo,
                 modo_comparacao, atualizado_em)
            VALUES (:codigo_indicador,:nome_indicador,:usa_dados_operacionais,
                    :aba_origem_excel,:campo_origem,:resultado_representa,
                    :subindicadores_existem,:subindicadores_status,
                    :status_mapeamento,:observacoes,:indicador_ativo,
                    :modo_comparacao,:atualizado_em)
            ON CONFLICT(codigo_indicador) DO UPDATE SET
                nome_indicador          = excluded.nome_indicador,
                usa_dados_operacionais  = excluded.usa_dados_operacionais,
                aba_origem_excel        = excluded.aba_origem_excel,
                campo_origem            = excluded.campo_origem,
                resultado_representa    = excluded.resultado_representa,
                subindicadores_existem  = excluded.subindicadores_existem,
                subindicadores_status   = excluded.subindicadores_status,
                status_mapeamento       = excluded.status_mapeamento,
                observacoes             = excluded.observacoes,
                indicador_ativo         = excluded.indicador_ativo,
                modo_comparacao         = excluded.modo_comparacao,
                atualizado_em           = excluded.atualizado_em
        """, record)
        conn.commit()
        return True
    except Exception as e:
        print(f"[DB] Erro ao salvar: {e}")
        return False
    finally:
        conn.close()


def delete_by_codigo(codigo: str) -> bool:
    conn = _connect()
    try:
        conn.execute("DELETE FROM indicadores_mapeamento WHERE codigo_indicador = ?", (codigo,))
        conn.commit()
        return True
    except Exception:
        return False
    finally:
        conn.close()


def import_from_excel(excel_path: str) -> tuple[int, str]:
    """
    Importa indicadores da aba Base_Indicadores do Excel.
    Retorna (quantidade_importada, mensagem).
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if "Base_Indicadores" not in wb.sheetnames:
            return 0, "Aba 'Base_Indicadores' não encontrada no arquivo."

        ws = wb["Base_Indicadores"]
        headers = [c.value for c in ws[1]]
        conn = _connect()
        count = 0
        now   = datetime.now().isoformat(sep=" ", timespec="seconds")
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                d = dict(zip(headers, row))
                codigo = d.get("codigo") or d.get("codigo_indicador") or d.get("Código") or row[0]
                nome   = d.get("nome") or d.get("nome_indicador") or d.get("Nome") or ""
                conn.execute("""
                    INSERT INTO indicadores_mapeamento
                        (codigo_indicador, nome_indicador, atualizado_em)
                    VALUES (?,?,?)
                    ON CONFLICT(codigo_indicador) DO UPDATE SET
                        nome_indicador = excluded.nome_indicador,
                        atualizado_em  = excluded.atualizado_em
                """, (str(codigo).strip(), str(nome).strip(), now))
                count += 1
            conn.commit()
            # Log
            conn.execute("""
                INSERT INTO importacoes_log (data_hora, arquivo, registros, status, mensagem)
                VALUES (?,?,?,'OK','Importação concluída com sucesso.')
            """, (now, os.path.basename(excel_path), count))
            conn.commit()
            return count, f"{count} registros importados com sucesso."
        finally:
            conn.close()
    except Exception as e:
        return 0, f"Erro na importação: {e}"


def get_stats() -> dict:
    """KPI cards da tela Base de Dados."""
    from mapping_db import STATUS_MAPEADO, STATUS_PENDENTE_PROCESSO, STATUS_PENDENTE_CONTROLE
    rows = get_all()
    total     = len(rows)
    mapeados  = sum(1 for r in rows if r["status_mapeamento"] == STATUS_MAPEADO)
    sem_vinc  = sum(1 for r in rows if r["status_mapeamento"] == "Sem vínculo")
    pendentes = sum(1 for r in rows if r["status_mapeamento"] in (
        STATUS_PENDENTE_PROCESSO, STATUS_PENDENTE_CONTROLE))
    conn = _connect()
    try:
        ultima = conn.execute(
            "SELECT data_hora, arquivo FROM importacoes_log ORDER BY id DESC LIMIT 1"
        ).fetchone()
        ultima_str = f"{ultima['data_hora']}\nArquivo: {ultima['arquivo']}" if ultima else "—"
    finally:
        conn.close()

    return {
        "total":        total,
        "mapeados":     mapeados,
        "pct_mapeados": round(mapeados / total * 100) if total else 0,
        "sem_vinculo":  sem_vinc,
        "pct_sem":      round(sem_vinc / total * 100) if total else 0,
        "pendentes":    pendentes,
        "pct_pend":     round(pendentes / total * 100) if total else 0,
        "linhas_banco": total,
        "ultima_import":ultima_str,
    }


# Inicializar banco e semear ao importar o módulo
init_db()
if is_empty():
    seed_from_mapping()
