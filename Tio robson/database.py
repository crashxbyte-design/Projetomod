"""
database.py - Backend SQLite persistente do sistema.
Arquivo: C:/Users/Admin/Documents/Robson/sp_indicadores.db

Tabelas:
  indicadores_mapeamento  — definição e mapeamento de cada indicador
  dados_historicos        — valores mensais reais lidos do Excel (ETL)
  config                  — chave/valor de configurações globais
  analise_critica         — registros de análise crítica por indicador/período
  importacoes_log         — histórico de importações do Excel
"""

import sqlite3
import os
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.normpath(os.path.join(BASE_DIR, "..", "sp_indicadores.db"))

# ── Schema ─────────────────────────────────────────────────────────────────

_SCHEMA_STATEMENTS = [
    # Tabela principal: definição e mapeamento dos indicadores
    """
    CREATE TABLE IF NOT EXISTS indicadores_mapeamento (
        id                       INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_indicador         TEXT    UNIQUE NOT NULL,
        nome_indicador           TEXT    NOT NULL,
        tipo                     TEXT    DEFAULT 'Operacional',
        periodicidade            TEXT    DEFAULT 'Mensal',
        unidade                  TEXT,
        meta_texto               TEXT,
        meta_numero              REAL,
        menor_melhor             INTEGER DEFAULT 1,
        usa_dados_operacionais   INTEGER DEFAULT 1,
        aba_origem_excel         TEXT,
        campo_origem             TEXT,
        resultado_representa     TEXT,
        subindicadores_existem   INTEGER DEFAULT 0,
        subindicadores_status    TEXT    DEFAULT 'A definir',
        status_mapeamento        TEXT    DEFAULT 'Sem vínculo',
        observacoes              TEXT,
        indicador_ativo          INTEGER DEFAULT 1,
        modo_comparacao          TEXT    DEFAULT '2025 x 2026',
        atualizado_em            TEXT
    )
    """,

    # Valores mensais reais, importados do Excel via ETL
    """
    CREATE TABLE IF NOT EXISTS dados_historicos (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_indicador TEXT    NOT NULL,
        ano              INTEGER NOT NULL,
        mes              TEXT    NOT NULL,
        valor            REAL,
        UNIQUE(codigo_indicador, ano, mes)
    )
    """,

    # Configurações globais (chave/valor)
    """
    CREATE TABLE IF NOT EXISTS config (
        chave TEXT PRIMARY KEY,
        valor TEXT
    )
    """,

    # Análise crítica por indicador
    """
    CREATE TABLE IF NOT EXISTS analise_critica (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_indicador TEXT    NOT NULL,
        periodo          TEXT,
        nivel            TEXT    DEFAULT 'ATENÇÃO',
        analise          TEXT,
        prazo            TEXT,
        atualizado_em    TEXT
    )
    """,

    # Log de importações
    """
    CREATE TABLE IF NOT EXISTS importacoes_log (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        data_hora  TEXT,
        arquivo    TEXT,
        registros  INTEGER,
        status     TEXT,
        mensagem   TEXT
    )
    """,
]

# Colunas a adicionar em versões futuras (migration segura)
_MIGRATIONS = [
    "ALTER TABLE indicadores_mapeamento ADD COLUMN tipo TEXT DEFAULT 'Operacional'",
    "ALTER TABLE indicadores_mapeamento ADD COLUMN periodicidade TEXT DEFAULT 'Mensal'",
    "ALTER TABLE indicadores_mapeamento ADD COLUMN unidade TEXT",
    "ALTER TABLE indicadores_mapeamento ADD COLUMN meta_texto TEXT",
    "ALTER TABLE indicadores_mapeamento ADD COLUMN meta_numero REAL",
    "ALTER TABLE indicadores_mapeamento ADD COLUMN menor_melhor INTEGER DEFAULT 1",
]


def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_db():
    """Cria tabelas e aplica migrations seguras."""
    conn = _connect()
    try:
        for stmt in _SCHEMA_STATEMENTS:
            conn.execute(stmt.strip())
        conn.commit()
        # Migrations: ADD COLUMN ignorando erros de "duplicate column"
        for mig in _MIGRATIONS:
            try:
                conn.execute(mig)
                conn.commit()
            except sqlite3.OperationalError:
                pass  # Coluna já existe
    finally:
        conn.close()


def is_empty() -> bool:
    conn = _connect()
    try:
        row = conn.execute("SELECT COUNT(*) FROM indicadores_mapeamento").fetchone()
        return row[0] == 0
    finally:
        conn.close()


def seed_from_mapping():
    """Bootstrap: popula indicadores_mapeamento e config a partir de mapping_db.py.
    Executado apenas uma vez quando o banco está vazio.
    NÃO é chamado durante o runtime normal do app.
    """
    from mapping_db import MAPEAMENTO_INDICADORES
    conn = _connect()
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    try:
        for m in MAPEAMENTO_INDICADORES:
            conn.execute("""
                INSERT OR IGNORE INTO indicadores_mapeamento
                    (codigo_indicador, nome_indicador, usa_dados_operacionais,
                     aba_origem_excel, campo_origem, resultado_representa,
                     subindicadores_existem, subindicadores_status,
                     status_mapeamento, observacoes, atualizado_em)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, (
                m["codigo_indicador"], m["nome_indicador"],
                1 if m["usa_dados_operacionais"] else 0,
                m.get("aba_origem_excel"), m.get("campo_origem"),
                m.get("resultado_representa"),
                1 if m.get("subindicadores_existem") else 0,
                m.get("subindicadores_status", "A definir"),
                m.get("status_mapeamento", "Sem vínculo"),
                m.get("observacoes"), now,
            ))
        # Config padrão
        defaults = [
            ("periodo_atual",    "Jan a Fev/2026"),
            ("responsavel_geral","Segurança Patrimonial"),
            ("data_atualizacao", now),
        ]
        for chave, valor in defaults:
            conn.execute(
                "INSERT OR IGNORE INTO config (chave, valor) VALUES (?,?)",
                (chave, valor)
            )
        conn.commit()
    finally:
        conn.close()


# ── CRUD: indicadores_mapeamento ───────────────────────────────────────────

def get_all() -> list[dict]:
    conn = _connect()
    try:
        rows = conn.execute(
            "SELECT * FROM indicadores_mapeamento ORDER BY codigo_indicador"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def get_active() -> list[dict]:
    """Retorna apenas indicadores ativos (indicador_ativo=1)."""
    conn = _connect()
    try:
        rows = conn.execute(
            "SELECT * FROM indicadores_mapeamento WHERE indicador_ativo=1 ORDER BY codigo_indicador"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def get_by_codigo(codigo: str) -> dict | None:
    conn = _connect()
    try:
        row = conn.execute(
            "SELECT * FROM indicadores_mapeamento WHERE codigo_indicador = ?", (codigo,)
        ).fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


def upsert(record: dict) -> bool:
    """Insere ou atualiza um registro de indicador. Retorna True se bem-sucedido."""
    conn = _connect()
    try:
        record = dict(record)  # cópia para não mutar o original
        record["atualizado_em"] = datetime.now().isoformat(sep=" ", timespec="seconds")
        # Garante que chaves ausentes não explodam o INSERT
        record.setdefault("tipo", "Operacional")
        record.setdefault("periodicidade", "Mensal")
        record.setdefault("unidade", None)
        record.setdefault("meta_texto", None)
        record.setdefault("meta_numero", None)
        record.setdefault("menor_melhor", 1)
        record.setdefault("subindicadores_status", "A definir")
        record.setdefault("indicador_ativo", 1)
        record.setdefault("modo_comparacao", "2025 x 2026")

        conn.execute("""
            INSERT INTO indicadores_mapeamento
                (codigo_indicador, nome_indicador, tipo, periodicidade, unidade,
                 meta_texto, meta_numero, menor_melhor,
                 usa_dados_operacionais, aba_origem_excel, campo_origem,
                 resultado_representa, subindicadores_existem, subindicadores_status,
                 status_mapeamento, observacoes, indicador_ativo,
                 modo_comparacao, atualizado_em)
            VALUES
                (:codigo_indicador,:nome_indicador,:tipo,:periodicidade,:unidade,
                 :meta_texto,:meta_numero,:menor_melhor,
                 :usa_dados_operacionais,:aba_origem_excel,:campo_origem,
                 :resultado_representa,:subindicadores_existem,:subindicadores_status,
                 :status_mapeamento,:observacoes,:indicador_ativo,
                 :modo_comparacao,:atualizado_em)
            ON CONFLICT(codigo_indicador) DO UPDATE SET
                nome_indicador         = excluded.nome_indicador,
                tipo                   = excluded.tipo,
                periodicidade          = excluded.periodicidade,
                unidade                = excluded.unidade,
                meta_texto             = excluded.meta_texto,
                meta_numero            = excluded.meta_numero,
                menor_melhor           = excluded.menor_melhor,
                usa_dados_operacionais = excluded.usa_dados_operacionais,
                aba_origem_excel       = excluded.aba_origem_excel,
                campo_origem           = excluded.campo_origem,
                resultado_representa   = excluded.resultado_representa,
                subindicadores_existem = excluded.subindicadores_existem,
                subindicadores_status  = excluded.subindicadores_status,
                status_mapeamento      = excluded.status_mapeamento,
                observacoes            = excluded.observacoes,
                indicador_ativo        = excluded.indicador_ativo,
                modo_comparacao        = excluded.modo_comparacao,
                atualizado_em          = excluded.atualizado_em
        """, record)
        conn.commit()
        return True
    except Exception as e:
        print(f"[DB] Erro ao salvar indicador: {e}")
        return False
    finally:
        conn.close()


def delete_by_codigo(codigo: str) -> bool:
    conn = _connect()
    try:
        conn.execute("DELETE FROM indicadores_mapeamento WHERE codigo_indicador = ?", (codigo,))
        conn.execute("DELETE FROM dados_historicos WHERE codigo_indicador = ?", (codigo,))
        conn.commit()
        return True
    except Exception:
        return False
    finally:
        conn.close()


# ── CRUD: dados_historicos ─────────────────────────────────────────────────

def upsert_historico(codigo: str, ano: int, mes: str, valor) -> bool:
    """Grava ou atualiza um valor mensal para um indicador."""
    conn = _connect()
    try:
        conn.execute("""
            INSERT INTO dados_historicos (codigo_indicador, ano, mes, valor)
            VALUES (?,?,?,?)
            ON CONFLICT(codigo_indicador, ano, mes) DO UPDATE SET valor=excluded.valor
        """, (codigo, ano, mes, valor))
        conn.commit()
        return True
    except Exception as e:
        print(f"[DB] Erro ao salvar histórico {codigo}/{ano}/{mes}: {e}")
        return False
    finally:
        conn.close()


def get_historico(codigo: str, anos: list[int] | None = None) -> dict:
    """Retorna {ano: {mes: valor}} para um indicador."""
    conn = _connect()
    try:
        if anos:
            placeholders = ",".join("?" * len(anos))
            rows = conn.execute(
                f"SELECT ano, mes, valor FROM dados_historicos "
                f"WHERE codigo_indicador=? AND ano IN ({placeholders}) ORDER BY ano, mes",
                [codigo] + anos
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT ano, mes, valor FROM dados_historicos WHERE codigo_indicador=? ORDER BY ano, mes",
                (codigo,)
            ).fetchall()
        resultado: dict = {}
        for r in rows:
            resultado.setdefault(r["ano"], {})[r["mes"]] = r["valor"]
        return resultado
    finally:
        conn.close()


def get_historico_multi(codigos: list[str], anos: list[int]) -> dict:
    """Retorna {codigo: {ano: {mes: valor}}} para múltiplos indicadores."""
    if not codigos or not anos:
        return {}
    conn = _connect()
    try:
        p_cod = ",".join("?" * len(codigos))
        p_ano = ",".join("?" * len(anos))
        rows = conn.execute(
            f"SELECT codigo_indicador, ano, mes, valor FROM dados_historicos "
            f"WHERE codigo_indicador IN ({p_cod}) AND ano IN ({p_ano}) ORDER BY ano",
            codigos + anos
        ).fetchall()
        resultado: dict = {}
        for r in rows:
            resultado.setdefault(r["codigo_indicador"], {}).setdefault(r["ano"], {})[r["mes"]] = r["valor"]
        return resultado
    finally:
        conn.close()


# ── CRUD: config ───────────────────────────────────────────────────────────

def get_config(chave: str, default: str = "") -> str:
    conn = _connect()
    try:
        row = conn.execute("SELECT valor FROM config WHERE chave=?", (chave,)).fetchone()
        return row["valor"] if row else default
    finally:
        conn.close()


def set_config(chave: str, valor: str):
    conn = _connect()
    try:
        conn.execute(
            "INSERT INTO config (chave, valor) VALUES (?,?) ON CONFLICT(chave) DO UPDATE SET valor=excluded.valor",
            (chave, valor)
        )
        conn.commit()
    finally:
        conn.close()


def get_all_config() -> dict:
    conn = _connect()
    try:
        rows = conn.execute("SELECT chave, valor FROM config").fetchall()
        return {r["chave"]: r["valor"] for r in rows}
    finally:
        conn.close()


# ── CRUD: analise_critica ──────────────────────────────────────────────────

def get_analise_critica(codigo: str | None = None) -> list[dict]:
    conn = _connect()
    try:
        if codigo:
            rows = conn.execute(
                "SELECT * FROM analise_critica WHERE codigo_indicador=? ORDER BY id",
                (codigo,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM analise_critica ORDER BY codigo_indicador, id"
            ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def upsert_analise_critica(codigo: str, analise: str, prazo: str = "", nivel: str = "ATENÇÃO", periodo: str = "") -> bool:
    conn = _connect()
    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    try:
        conn.execute("""
            INSERT INTO analise_critica (codigo_indicador, periodo, nivel, analise, prazo, atualizado_em)
            VALUES (?,?,?,?,?,?)
        """, (codigo, periodo, nivel, analise, prazo, now))
        conn.commit()
        return True
    except Exception as e:
        print(f"[DB] Erro ao salvar análise: {e}")
        return False
    finally:
        conn.close()


# ── Importação do Excel (ETL completo) ────────────────────────────────────

MESES = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]

# Mapeamento: aba do Excel → código do indicador
# É usado durante a importação para vincular automaticamente os dados históricos
# ao indicador correto com base na aba de origem configurada no banco.
# Se o usuário mapear a "Aba (Origem)" na tela Base de Dados,
# a importação funcionará automaticamente.

def _read_sheet_generic(ws) -> dict:
    """
    Leitura genérica: busca coluna de meses e colunas de anos.
    Retorna {ano: {mes: valor}}.
    """
    anos_map = {}  # {ano: col_index}
    mes_col = None
    resultado: dict = {}

    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        for i, v in enumerate(row):
            if isinstance(v, int) and 2020 <= v <= 2030:
                anos_map[v] = i
            if isinstance(v, str) and v.strip() in MESES and mes_col is None:
                mes_col = i

    if mes_col is None:
        # Tenta detectar mês na coluna 0 ou 1
        for row in ws.iter_rows(min_row=2, max_row=20, values_only=True):
            for ci in [0, 1]:
                val = str(row[ci]).strip() if row[ci] else ""
                if val in MESES:
                    mes_col = ci
                    break
            if mes_col is not None:
                break

    if mes_col is None or not anos_map:
        return resultado

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or mes_col >= len(row):
            continue
        mes = str(row[mes_col]).strip() if row[mes_col] else ""
        if mes not in MESES:
            continue
        for ano, ci in anos_map.items():
            if ci < len(row):
                val = row[ci]
                if isinstance(val, (int, float)):
                    resultado.setdefault(ano, {})[mes] = val

    return resultado


def _import_sheet_data(conn, codigo: str, ws):
    """Lê uma aba do Excel e persiste em dados_historicos para o indicador dado."""
    dados = _read_sheet_generic(ws)
    count = 0
    for ano, meses_dict in dados.items():
        for mes, valor in meses_dict.items():
            conn.execute("""
                INSERT INTO dados_historicos (codigo_indicador, ano, mes, valor)
                VALUES (?,?,?,?)
                ON CONFLICT(codigo_indicador, ano, mes) DO UPDATE SET valor=excluded.valor
            """, (codigo, ano, mes, valor))
            count += 1
    return count


def import_from_excel(excel_path: str) -> tuple[int, str]:
    """
    ETL completo do Excel para o SQLite.

    Fluxo:
      1. Lê a aba 'Base_Indicadores' (se existir) para atualizar nomes/metadados.
      2. Para cada indicador ativo no banco com 'aba_origem_excel' mapeado,
         localiza a aba no arquivo e importa os dados mensais para dados_historicos.
      3. Lê 'Config' do Excel (se existir) e atualiza a tabela config.
      4. Lê 'Base_Analise_Critica' (se existir) e importa registros.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        return 0, f"Erro ao abrir arquivo: {e}"

    now = datetime.now().isoformat(sep=" ", timespec="seconds")
    total_registros = 0
    conn = _connect()

    try:
        # 1. Config
        if "Config" in wb.sheetnames:
            for row in wb["Config"].iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] is not None:
                    conn.execute(
                        "INSERT INTO config (chave, valor) VALUES (?,?) "
                        "ON CONFLICT(chave) DO UPDATE SET valor=excluded.valor",
                        (str(row[0]).strip(), str(row[1]).strip())
                    )

        # 2. Base_Indicadores
        if "Base_Indicadores" in wb.sheetnames:
            ws = wb["Base_Indicadores"]
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                d = dict(zip(headers, row))
                codigo = str(d.get("codigo") or d.get("codigo_indicador") or row[0]).strip()
                nome = str(d.get("nome") or d.get("nome_indicador") or "").strip()
                if not codigo:
                    continue
                conn.execute("""
                    INSERT INTO indicadores_mapeamento (codigo_indicador, nome_indicador, atualizado_em)
                    VALUES (?,?,?)
                    ON CONFLICT(codigo_indicador) DO UPDATE SET
                        nome_indicador=excluded.nome_indicador,
                        atualizado_em=excluded.atualizado_em
                """, (codigo, nome, now))
                total_registros += 1

        # 3. Dados históricos: para cada indicador mapeado no banco, tenta importar a aba
        indicadores = conn.execute(
            "SELECT codigo_indicador, aba_origem_excel FROM indicadores_mapeamento "
            "WHERE indicador_ativo=1 AND aba_origem_excel IS NOT NULL AND aba_origem_excel != ''"
        ).fetchall()

        for ind in indicadores:
            aba = ind["aba_origem_excel"]
            cod = ind["codigo_indicador"]

            # Procura a aba no workbook (match exato ou parcial case-insensitive)
            ws_target = None
            for sname in wb.sheetnames:
                if sname.upper() == aba.upper() or aba.upper() in sname.upper():
                    ws_target = wb[sname]
                    break

            if ws_target is None:
                continue

            n = _import_sheet_data(conn, cod, ws_target)
            total_registros += n

            # Marca status como Mapeado se importou dados
            if n > 0:
                conn.execute(
                    "UPDATE indicadores_mapeamento SET status_mapeamento='Mapeado', atualizado_em=? "
                    "WHERE codigo_indicador=?", (now, cod)
                )

        # 4. Base_Analise_Critica
        if "Base_Analise_Critica" in wb.sheetnames:
            ws = wb["Base_Analise_Critica"]
            headers_ac = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                d = dict(zip(headers_ac, row))
                cod_ac = str(d.get("codigo_indicador") or row[0]).strip()
                analise = str(d.get("analise_critica") or d.get("analise") or "").strip()
                prazo = str(d.get("prazo") or "").strip()
                if cod_ac and analise:
                    conn.execute("""
                        INSERT INTO analise_critica (codigo_indicador, analise, prazo, atualizado_em)
                        VALUES (?,?,?,?)
                    """, (cod_ac, analise, prazo, now))
                    total_registros += 1

        conn.commit()

        # Log
        conn.execute("""
            INSERT INTO importacoes_log (data_hora, arquivo, registros, status, mensagem)
            VALUES (?,?,?,'OK','ETL concluído com sucesso.')
        """, (now, os.path.basename(excel_path), total_registros))
        conn.commit()

        return total_registros, f"{total_registros} registros importados/atualizados de '{os.path.basename(excel_path)}'."

    except Exception as e:
        conn.rollback()
        return 0, f"Erro durante ETL: {e}"
    finally:
        conn.close()


# ── Stats (KPIs da tela Base de Dados) ────────────────────────────────────

def get_stats() -> dict:
    rows = get_all()
    total     = len(rows)
    mapeados  = sum(1 for r in rows if r["status_mapeamento"] == "Mapeado")
    sem_vinc  = sum(1 for r in rows if r["status_mapeamento"] == "Sem vínculo")
    pendentes = sum(1 for r in rows if r["status_mapeamento"] in (
        "Pendente de processo", "Pendente de controle"))
    conn = _connect()
    try:
        ultima = conn.execute(
            "SELECT data_hora, arquivo FROM importacoes_log ORDER BY id DESC LIMIT 1"
        ).fetchone()
        ultima_str = f"{ultima['data_hora']}\nArquivo: {ultima['arquivo']}" if ultima else "—"
    finally:
        conn.close()

    return {
        "total":        total,
        "mapeados":     mapeados,
        "pct_mapeados": round(mapeados / total * 100) if total else 0,
        "sem_vinculo":  sem_vinc,
        "pct_sem":      round(sem_vinc / total * 100) if total else 0,
        "pendentes":    pendentes,
        "pct_pend":     round(pendentes / total * 100) if total else 0,
        "linhas_banco": total,
        "ultima_import": ultima_str,
    }


# ── Bootstrap ─────────────────────────────────────────────────────────────

init_db()
if is_empty():
    seed_from_mapping()
