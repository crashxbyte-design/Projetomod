"""
data_loader.py - Camada de dados do Dashboard de Segurança Patrimonial.

Fontes:
  - INDICADORES CONTROLE GERAL SEGURANÇA - 2026.xlsx  → dados históricos reais (2024/2025/2026)
  - BD_Dashboard_Seguranca.xlsx                        → Config, Analise_Critica, campos auxiliares
"""

import os
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

PLANILHA_OPERACIONAL = os.path.join(
    BASE_DIR, "INDICADORES CONTROLE GERAL SEGURANÇA - 2026.xlsx"
)

MESES = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]


def _find_mes_col(ws, meses=MESES):
    """Retorna dicionário {nome_mes: row_index} baseado na coluna de meses da planilha."""
    mes_rows = {}
    for row in ws.iter_rows():
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            for m in meses:
                if val == m or val == m[:3]:
                    mes_rows[m] = cell.row
                    break
    return mes_rows


def _load_bd():
    """Carrega Config, Base_Analise_Critica e Base_Indicadores da planilha única."""
    wb = openpyxl.load_workbook(PLANILHA_OPERACIONAL, data_only=True)

    config = {}
    for row in wb["Config"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            config[row[0]] = row[1]

    ac_raw = []
    headers = [c.value for c in wb["Base_Analise_Critica"][1]]
    for row in wb["Base_Analise_Critica"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            ac_raw.append(dict(zip(headers, row)))

    indicadores_bd = []
    if "Base_Indicadores" in wb.sheetnames:
        headers_ind = [c.value for c in wb["Base_Indicadores"][1]]
        for row in wb["Base_Indicadores"].iter_rows(min_row=2, values_only=True):
            if row[0]:
                indicadores_bd.append(dict(zip(headers_ind, row)))

    return config, ac_raw, indicadores_bd


# ---------------------------------------------------------------------------
# Leitores por aba da planilha operacional
# ---------------------------------------------------------------------------

def _read_evasao(wb):
    """Aba EVASÃO: colunas anos 2021..2026 em col 2..7 a partir da linha 3."""
    ws = wb.sheetnames[0]  # EVASÃO
    ws = wb[wb.sheetnames[0]]
    anos = {}
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        if row[2] in (2021, 2022, 2023, 2024, 2025, 2026):
            header_row = [row[i] for i in range(1, 8)]  # Meses + anos
            anos = {row[i]: i for i in range(2, 8) if isinstance(row[i], int)}
            break

    resultado = {}  # {ano: {mes: valor}}
    for row in ws.iter_rows(min_row=5, values_only=True):
        mes = str(row[1]).strip() if row[1] else None
        if mes in MESES:
            for ano, col_idx in anos.items():
                val = row[col_idx]
                resultado.setdefault(ano, {})[mes] = val
    return resultado


def _read_morgue(wb):
    """Aba MORGUE: Óbitos. Anos 2022..2026 em cols 2..6."""
    ws = wb["MORGUE"]
    anos = {}
    for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
        if isinstance(row[2], int) and row[2] >= 2022:
            for i, v in enumerate(row[1:], 1):
                if isinstance(v, int): anos[v] = i
            break
    resultado = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        mes = str(row[1]).strip() if row[1] else None
        if mes in MESES:
            for ano, col_idx in anos.items():
                val = row[col_idx] if col_idx < len(row) else None
                resultado.setdefault(ano, {})[mes] = val
    return resultado


def _read_patio(wb):
    """Aba PATIO: Total A+B colunas 4, 7, 10 para 2024, 2025, 2026."""
    ws = wb["PATIO"]
    resultado = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        mes = str(row[1]).strip() if row[1] else None
        if mes in MESES:
            resultado.setdefault(2024, {})[mes] = row[4]   # Total A+B 2024
            resultado.setdefault(2025, {})[mes] = row[7]   # Total A+B 2025
            resultado.setdefault(2026, {})[mes] = row[10]  # Total A+B 2026
    return resultado


def _read_patio_nutri(wb):
    """Aba PATIO NUTRI: 2025 col 1, 2026 col 2."""
    ws = wb["PATIO NUTRI"]
    resultado = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        mes = str(row[0]).strip() if row[0] else None
        if mes in MESES:
            resultado.setdefault(2025, {})[mes] = row[1]
            resultado.setdefault(2026, {})[mes] = row[2]
    return resultado


def _read_apreensoes(wb):
    """Aba APREENSÕES: Armas+Tabaco/Entorpecentes agrupados como total."""
    ws_name = next(n for n in wb.sheetnames if "PREEN" in n.upper())
    ws = wb[ws_name]
    # cols: Armas 2024=3, 2025=4, 2026=5 | Tabaco 2024=7, 2025=8, 2026=9
    resultado = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        mes = str(row[2]).strip() if row[2] else None
        if mes in MESES:
            for ano, ca, ct in [(2024, 3, 7), (2025, 4, 8), (2026, 5, 9)]:
                v_a = row[ca] if ca < len(row) else None
                v_t = row[ct] if ct < len(row) else None
                total = (v_a or 0) + (v_t or 0) if (v_a is not None or v_t is not None) else None
                resultado.setdefault(ano, {})[mes] = total
    return resultado


def _read_intercorrencias(wb):
    """Aba INTERCORRENCIAS SEG: coluna INTERCORRÊNCIAS Total 2024/2025/2026 (cols 10/11/12)."""
    ws_name = next(n for n in wb.sheetnames if "INTERCOR" in n.upper())
    ws = wb[ws_name]
    resultado = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        mes = str(row[3]).strip() if row[3] else None
        if mes in MESES:
            resultado.setdefault(2024, {})[mes] = row[10]
            resultado.setdefault(2025, {})[mes] = row[11]
            resultado.setdefault(2026, {})[mes] = row[12]
    return resultado


def _read_visitas(wb):
    """Aba VISITAS: Visitantes + Acompanhantes por ano."""
    ws = wb["VISITAS"]
    # Visitantes: cols 2=2023, 3=2024, 4=2025, 5=2026
    # Acompanhantes: cols 9=2023, 10=2024, 11=2025, 12=2026
    resultado_vis = {}
    resultado_acomp = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        mes = str(row[1]).strip() if row[1] else None
        if mes in MESES:
            resultado_vis.setdefault(2024, {})[mes] = row[3]
            resultado_vis.setdefault(2025, {})[mes] = row[4]
            resultado_vis.setdefault(2026, {})[mes] = row[5]
            resultado_acomp.setdefault(2024, {})[mes] = row[10]
            resultado_acomp.setdefault(2025, {})[mes] = row[11]
            resultado_acomp.setdefault(2026, {})[mes] = row[12]
    return resultado_vis, resultado_acomp


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------

def get_all_data():
    wb_op = openpyxl.load_workbook(PLANILHA_OPERACIONAL, data_only=True)
    config, ac_raw, indicadores_bd = _load_bd()

    # Leitura de cada indicador operacional
    evasao       = _read_evasao(wb_op)
    morgue       = _read_morgue(wb_op)
    patio        = _read_patio(wb_op)
    patio_nutri  = _read_patio_nutri(wb_op)
    apreensoes   = _read_apreensoes(wb_op)
    intercorr    = _read_intercorrencias(wb_op)
    visitantes, acompanhantes = _read_visitas(wb_op)

    # Indicadores fixos do BD_Dashboard para fallback de metadata
    INDICADORES_META = {i['codigo']: i for i in indicadores_bd}

    def _status(cod, val_atual, meta_num=None, menor_melhor=True):
        """Calcula status com base no valor e meta numérica."""
        cfg = INDICADORES_META.get(cod, {})
        if cfg.get("status") and cfg["status"] != "Calculado":
            return cfg["status"]
        if val_atual is None:
            return "A preencher"
        if meta_num is None:
            return "Em Atenção"
        if menor_melhor:
            if val_atual <= meta_num: return "Dentro da meta"
            elif val_atual <= meta_num * 1.2: return "Em Atenção"
            else: return "Acima da meta"
        else:
            if val_atual >= meta_num: return "Dentro da meta"
            elif val_atual >= meta_num * 0.8: return "Em Atenção"
            else: return "Abaixo da meta"

    # Comparativos 2025 x 2026
    comparativos = {
        "SP.IND.006": {"nome": "Evasão de Pacientes", "unidade": "evasões", "dados": evasao, "meta_num": 3, "menor_melhor": True},
        "SP.IND.001": {"nome": "Óbitos Recolhidos (Morgue)", "unidade": "óbitos", "dados": morgue, "meta_num": 140, "menor_melhor": True},
        "SP.IND.005": {"nome": "Acesso ao Pátio (Total A+B)", "unidade": "acessos", "dados": patio, "meta_num": None, "menor_melhor": False},
        "SP.IND.007": {"nome": "Apreensões Totais", "unidade": "apreensões", "dados": apreensoes, "meta_num": None, "menor_melhor": True},
        "SP.IND.004_vis": {"nome": "Visitantes", "unidade": "visitas", "dados": visitantes, "meta_num": None, "menor_melhor": False},
        "SP.IND.004_acomp": {"nome": "Acompanhantes", "unidade": "acomp.", "dados": acompanhantes, "meta_num": None, "menor_melhor": False},
        "SP.IND.001_ic": {"nome": "Intercorrências de Segurança", "unidade": "intercorrências", "dados": intercorr, "meta_num": 140, "menor_melhor": True},
        "PATIO_NUTRI": {"nome": "Pátio Nutrição", "unidade": "entregas", "dados": patio_nutri, "meta_num": None, "menor_melhor": False},
    }

    def _get_meta(cod, default):
        return str(INDICADORES_META.get(cod, {}).get("meta") or default)

    def _get_titulo(cod, default):
        return INDICADORES_META.get(cod, {}).get("nome") or default

    indicadores = [
        {
            "codigo": "SP.IND.001",
            "titulo": _get_titulo("SP.IND.001", "Incidência de Intercorrências de Segurança Patrimonial"),
            "tipo": "Operacional", "periodicidade": "Mensal",
            "meta": _get_meta("SP.IND.001", "≤ 140"),
            "resultado_jan": intercorr.get(2026, {}).get("Janeiro"),
            "resultado_fev": intercorr.get(2026, {}).get("Fevereiro"),
            "status": _status("SP.IND.001", intercorr.get(2026, {}).get("Fevereiro") or intercorr.get(2026, {}).get("Janeiro"), 140),
            "origem": "Aba INTERCORRENCIAS SEG – Planilha Operacional",
        },
        {
            "codigo": "SP.IND.002",
            "titulo": _get_titulo("SP.IND.002", "Índice de Conformidade de Segurança Patrimonial"),
            "tipo": "Operacional", "periodicidade": "Mensal",
            "meta": _get_meta("SP.IND.002", "Meta a definir"),
            "resultado_jan": None, "resultado_fev": None,
            "status": _status("SP.IND.002", None),
            "origem": "Necessita implantação de processo + checklist de ronda",
        },
        {
            "codigo": "SP.IND.003",
            "titulo": _get_titulo("SP.IND.003", "Atendimento às Solicitações de Óbito"),
            "tipo": "Operacional", "periodicidade": "Mensal",
            "meta": _get_meta("SP.IND.003", "Meta a definir"),
            "resultado_jan": morgue.get(2026, {}).get("Janeiro"),
            "resultado_fev": morgue.get(2026, {}).get("Fevereiro"),
            "status": "Em Atenção" if morgue.get(2026, {}).get("Fevereiro") or morgue.get(2026, {}).get("Janeiro") else "A preencher",
            "origem": "Aba MORGUE – Planilha Operacional",
        },
        {
            "codigo": "SP.IND.004",
            "titulo": _get_titulo("SP.IND.004", "Acesso de Pessoas à Instituição"),
            "tipo": "Operacional", "periodicidade": "Mensal",
            "meta": _get_meta("SP.IND.004", "Meta a definir"),
            "resultado_jan": (visitantes.get(2026, {}).get("Janeiro") or 0) + (acompanhantes.get(2026, {}).get("Janeiro") or 0) or None,
            "resultado_fev": (visitantes.get(2026, {}).get("Fevereiro") or 0) + (acompanhantes.get(2026, {}).get("Fevereiro") or 0) or None,
            "status": _status("SP.IND.004", None), # Will fallback to Em Atenção or A preencher or whatever is in excel
            "origem": "Aba VISITAS – Planilha Operacional",
        },
        {
            "codigo": "SP.IND.005",
            "titulo": _get_titulo("SP.IND.005", "Acesso de Veículos"),
            "tipo": "Operacional", "periodicidade": "Mensal",
            "meta": _get_meta("SP.IND.005", "Meta a definir"),
            "resultado_jan": patio.get(2026, {}).get("Janeiro"),
            "resultado_fev": patio.get(2026, {}).get("Fevereiro"),
            "status": _status("SP.IND.005", None),
            "origem": "Aba PATIO – Planilha Operacional",
        },
        {
            "codigo": "SP.IND.006",
            "titulo": _get_titulo("SP.IND.006", "Evasão de Pacientes"),
            "tipo": "Operacional", "periodicidade": "Mensal",
            "meta": _get_meta("SP.IND.006", "≤ 3"),
            "resultado_jan": evasao.get(2026, {}).get("Janeiro"),
            "resultado_fev": evasao.get(2026, {}).get("Fevereiro"),
            "status": _status("SP.IND.006", evasao.get(2026, {}).get("Fevereiro") or evasao.get(2026, {}).get("Janeiro"), 3),
            "origem": "Aba EVASÃO – Planilha Operacional",
        },
        {
            "codigo": "SP.IND.007",
            "titulo": _get_titulo("SP.IND.007", "Apreensões"),
            "tipo": "Operacional", "periodicidade": "Mensal",
            "meta": _get_meta("SP.IND.007", "Meta a definir"),
            "resultado_jan": apreensoes.get(2026, {}).get("Janeiro"),
            "resultado_fev": apreensoes.get(2026, {}).get("Fevereiro"),
            "status": _status("SP.IND.007", None),
            "origem": "Aba APREENSÕES – Planilha Operacional",
        },
        {
            "codigo": "SP.IND.008",
            "titulo": _get_titulo("SP.IND.008", "Pacientes Monitorados por Dispositivos Tornozelados"),
            "tipo": "Operacional", "periodicidade": "Mensal",
            "meta": _get_meta("SP.IND.008", "Meta a definir"),
            "resultado_jan": None, "resultado_fev": None,
            "status": _status("SP.IND.008", None),
            "origem": "Necessita criação de controle formal com registro e periodicidade",
        },
        {
            "codigo": "SP.IND.009",
            "titulo": _get_titulo("SP.IND.009", "Qualidade dos Fornecedores de Segurança Patrimonial"),
            "tipo": "Estratégico", "periodicidade": "Semestral",
            "meta": _get_meta("SP.IND.009", "Meta a definir"),
            "resultado_jan": None, "resultado_fev": None,
            "status": _status("SP.IND.009", None),
            "origem": "Planilha controlada pela Qualidade",
        },
        {
            "codigo": "SP.IND.010",
            "titulo": _get_titulo("SP.IND.010", "Variação Orçamentária do Setor"),
            "tipo": "Estratégico", "periodicidade": "Mensal",
            "meta": _get_meta("SP.IND.010", "Meta a definir"),
            "resultado_jan": None, "resultado_fev": None,
            "status": _status("SP.IND.010", None),
            "origem": "A preencher",
        },
        {
            "codigo": "SP.IND.011",
            "titulo": _get_titulo("SP.IND.011", "Variação de Custos do Setor"),
            "tipo": "Estratégico", "periodicidade": "Mensal",
            "meta": _get_meta("SP.IND.011", "Meta a definir"),
            "resultado_jan": None, "resultado_fev": None,
            "status": _status("SP.IND.011", None),
            "origem": "A preencher",
        },
    ]

    # Filtrar para exibir apenas os indicadores que realmente existem na aba Base_Indicadores
    if INDICADORES_META:
        indicadores = [i for i in indicadores if i["codigo"] in INDICADORES_META]

    # Pendências
    pendencias = []
    for ac in ac_raw:
        cod = ac['codigo_indicador']
        ind = next((i for i in indicadores if i['codigo'] == cod), None)
        pendencias.append({
            "codigo": cod,
            "titulo": ind['titulo'] if ind else cod,
            "nivel": "CRÍTICO" if "002" in cod else "ATENÇÃO",
            "descricao": ac['analise_critica'],
            "prazo": ac['prazo'],
        })

    # Stats
    total = len(indicadores)
    com_meta = sum(1 for i in indicadores if "definir" not in str(i.get("meta", "")).lower())
    sem_meta = total - com_meta
    em_atencao = sum(1 for i in indicadores if i.get("status") in ("Em Atenção", "Acima da meta"))
    pendentes = sum(1 for i in indicadores if "Pendente" in str(i.get("status", "")))
    a_preencher = sum(1 for i in indicadores if i.get("status") == "A preencher")

    stats = {
        "total": total,
        "com_meta": com_meta,
        "sem_meta": sem_meta,
        "em_atencao": em_atencao,
        "pendentes_processo": pendentes,
        "a_preencher": a_preencher,
        "periodo": config.get("periodo_atual", "Jan a Fev/2026"),
        "responsavel": config.get("responsavel_geral", "Segurança Patrimonial"),
        "atualizacao": config.get("data_atualizacao", ""),
    }

    # Sub_raw no formato esperado pelos painéis (para sparklines e gráficos de subind.)
    sub_raw = []
    for cod_chave, info in comparativos.items():
        dados = info["dados"]
        for ano in [2025, 2026]:   # apenas 2025 e 2026
            for mes in MESES:
                val = dados.get(ano, {}).get(mes)
                if val is not None:
                    sub_raw.append({
                        "codigo_indicador": cod_chave,
                        "nome_subindicador": info["nome"],
                        "mes": mes,
                        "ano": ano,
                        "valor": val,
                        "meta": info.get("meta_num"),
                    })

    return {
        "config": config,
        "indicadores": indicadores,
        "pendencias": pendencias,
        "stats": stats,
        "comparativos": comparativos,
        "sub_raw": sub_raw,
        "ac_raw": ac_raw,
    }
